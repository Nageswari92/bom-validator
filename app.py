import io
import os
import zipfile
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# 1. Branding & Page Configurations
st.set_page_config(page_title="BAVA TECH - BOM Validator", layout="wide", page_icon="🚀")

# Credentials System (Custom Corporate Authentication)
VALID_USER = "Bavatech"
VALID_PASS = "Review@26"

# Session State Initialization (Security Pipeline Tracking)
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

# ==========================================
# 2. Secure Login Portal Layer
# ==========================================
if not st.session_state["logged_in"]:
    st.html("""
        <div style="background-color:#1e293b; padding:25px; border-radius:10px; margin-bottom:25px; text-align:center;">
            <h2 style="color:white; margin:0; font-family:sans-serif;">🔒 BAVA TECH Security Gateway</h2>
            <p style="color:#94a3b8; margin:5px 0 0 0; font-size:14px;">Authorized Personnel Only — Corporate Audit Pipeline</p>
        </div>
    """)
    
    # UI Framework for Centering the Login Box
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        with st.form("login_form"):
            st.subheader("Login to Access Matrix")
            username = st.text_input("Username", placeholder="Enter username...")
            password = st.text_input("Password", type="password", placeholder="Enter password...")
            submit_login = st.form_submit_button("Authenticate Engine 🚀")
            
            if submit_login:
                if username == VALID_USER and password == VALID_PASS:
                    st.session_state["logged_in"] = True
                    st.success("Authentication Successful! Loading Core Matrix...")
                    st.rerun()
                else:
                    st.error("Invalid Username or Password. Pipeline Access Denied!")
                    
else:
    # ==========================================
    # 3. Main BOM Matrix System (Authenticated)
    # ==========================================
    # Action Header Row with Logout Utility
    top_col1, top_col2 = st.columns([8, 2])
    with top_col2:
        if st.button("🔒 Logout from System", use_container_width=True):
            st.session_state["logged_in"] = False
            st.rerun()

    st.html("""
        <div style="background-color:#0f4c81; padding:20px; border-radius:10px; margin-bottom:25px;">
            <h1 style="color:white; margin:0; font-family:sans-serif;">🚀 BAVA TECH | Advanced BOM Audit Matrix</h1>
            <p style="color:#e0e0e0; margin:5px 0 0 0; font-size:14px;">Proprietary Automated Validation Engine — Memory Stream Pipeline</p>
        </div>
    """)

    st.write("Upload the main structural BOM target along with component data (ZIP folder or raw multi-selection .xlsx files).")

    # Double File Upload Section
    col1, col2 = st.columns(2)

    with col1:
        main_bom_file = st.file_uploader("1. Upload Main Indented BOM File", type=["xls", "xlsx"])

    with col2:
        manual_inputs = st.file_uploader(
            "2. Upload Component Data (ZIP Folder or Multiple .xlsx Files)", 
            type=["zip", "xlsx"], 
            accept_multiple_files=True
        )

    uploaded_files_map = {}

    if manual_inputs:
        for uploaded_item in manual_inputs:
            filename = uploaded_item.name
            
            if filename.lower().endswith('.zip'):
                try:
                    with zipfile.ZipFile(uploaded_item) as z:
                        for file_info in z.infolist():
                            if file_info.is_dir():
                                continue
                            base_name = os.path.basename(file_info.filename)
                            if base_name.endswith('.xlsx') and not base_name.startswith('~$'):
                                uploaded_files_map[base_name] = z.read(file_info.filename)
                except Exception as e:
                    st.error(f"Error reading ZIP file '{filename}': {str(e)}")
            
            elif filename.lower().endswith('.xlsx') and not filename.startswith('~$'):
                uploaded_files_map[filename] = uploaded_item.read()

    if main_bom_file is not None and uploaded_files_map:
        st.info(f"**BAVA TECH Core Engine:** Main BOM loaded. Found **{len(uploaded_files_map)}** component file(s) across input channels.")
        
        # Process BOM Data
        with st.spinner("Processing Source Data..."):
            df_all = pd.read_excel(main_bom_file, sheet_name=None, dtype=str)
            df_bom = df_all["BOM"] if "BOM" in df_all else list(df_all.values())[0]
            df_bom.columns = df_bom.columns.str.strip()
            df_bom['Level'] = df_bom['Level'].astype(int)
            df_bom['Part Number'] = df_bom['Part Number'].astype(str).str.strip()

            df_mfg = df_all["MFG"] if "MFG" in df_all else None
            mfg_lookup = {}
            if df_mfg is not None:
                df_mfg.columns = df_mfg.columns.str.strip()
                for idx, row in df_mfg.iterrows():
                    p_num = str(row['Part Number']).strip()
                    mfg_name = str(row['Manufacturer Name']).strip() if pd.notna(row['Manufacturer Name']) else ""
                    mfg_pn = str(row['Manufacturer Part Number']).strip() if pd.notna(row['Manufacturer Part Number']) else ""
                    
                    if p_num not in mfg_lookup:
                        mfg_lookup[p_num] = []
                    mfg_lookup[p_num].append({'mfg_name': mfg_name, 'mfg_pn': mfg_pn})

            # Hierarchy Tracking Logic
            assemblies = {}
            current_parent_at_lvl = {0: None}

            for idx, row in df_bom.iterrows():
                lvl = row['Level']
                part_no = row['Part Number']
                
                parent_part = current_parent_at_lvl.get(lvl - 1)
                if parent_part is not None:
                    assemblies[parent_part]['children'].append(row)
                    
                current_parent_at_lvl[lvl] = part_no
                
                if part_no not in assemblies:
                    assemblies[part_no] = {'info': row, 'children': []}
                if lvl < len(current_parent_at_lvl) - 1 or idx == 0:
                    assemblies[part_no]['info'] = row

        # Review and Validate Section
        if st.button("🚀 Run BAVA TECH Audit System"):
            unified_report_rows = []
            has_any_errors = False
            
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            progress_bar = st.progress(0)
            total_assemblies = len(assemblies)

            for index, (parent_part, data) in enumerate(assemblies.items()):
                progress_bar.progress((index + 1) / total_assemblies)
                
                child_rows = data['children']
                if not child_rows:
                    continue 
                    
                parent_info = data['info']
                parent_rev = str(parent_info['Revision']).strip() if pd.notna(parent_info['Revision']) else ""
                if parent_rev.lower() == 'nan':
                    parent_rev = ""
                elif len(parent_rev) == 1 and parent_rev.isdigit():
                    parent_rev = parent_rev.zfill(2)
                
                expected_filename = f"{parent_part}_REV_{parent_rev}_VSE_00_BOM.xlsx"
                
                if expected_filename not in uploaded_files_map:
                    has_any_errors = True
                    unified_report_rows.append({
                        "File": expected_filename, 
                        "Audit Result": "FAILED (Data Mismatch)", 
                        "Row": "N/A", 
                        "Field": "File Existence", "Expected": "File should be provided in inputs", 
                        "Found": "File Missing in Uploaded Data", "Status": "CRITICAL"
                    })
                    continue
                    
                local_file_errors = []
                try:
                    target_file_bytes = uploaded_files_map[expected_filename]
                    wb = load_workbook(io.BytesIO(target_file_bytes))
                    ws = wb["VSE_BOM"] if "VSE_BOM" in wb.sheetnames else wb.active
                    file_modified = [False]

                    def get_dynamic_format(orig_val_str, found_raw_val):
                        if '.' in orig_val_str:
                            decimal_places = len(orig_val_str.split('.')[1])
                            try:
                                return f"{float(found_raw_val):.{decimal_places}f}"
                            except:
                                pass
                        return str(found_raw_val if found_raw_val is not None else '').strip()

                    def check_cell(cell_coord, field_name, expected_val):
                        exp_val_str = str(expected_val).strip()
                        raw_found = ws[cell_coord].value
                        found_val = get_dynamic_format(exp_val_str, raw_found)

                        if found_val != exp_val_str:
                            ws[cell_coord].fill = yellow_fill
                            file_modified[0] = True
                            local_file_errors.append({
                                "Row": cell_coord, "Field": field_name,
                                "Expected": exp_val_str, "Found": found_val, "Status": "ERROR"
                            })

                    def check_mfg_cells(mfg_cell, pn_cell, part_num_key, row_label):
                        raw_mfg = ws[mfg_cell].value
                        raw_pn = ws[pn_cell].value

                        if part_num_key in mfg_lookup:
                            valid_options = mfg_lookup[part_num_key]
                            match_found = False
                            expected_display_list = []
                            
                            for option in valid_options:
                                exp_mfg = str(option['mfg_name']).strip()
                                exp_pn = str(option['mfg_pn']).strip()
                                
                                fmt_found_mfg = get_dynamic_format(exp_mfg, raw_mfg)
                                fmt_found_pn = get_dynamic_format(exp_pn, raw_pn)
                                
                                expected_display_list.append(f"({exp_mfg} / {exp_pn})")
                                
                                if fmt_found_mfg == exp_mfg and fmt_found_pn == exp_pn:
                                    match_found = True
                                    break
                            
                            if not match_found:
                                fallback_mfg = str(valid_options[0]['mfg_name']).strip()
                                fallback_pn = str(valid_options[0]['mfg_pn']).strip()
                                
                                final_found_mfg = get_dynamic_format(fallback_mfg, raw_mfg)
                                final_found_pn = get_dynamic_format(fallback_pn, raw_pn)

                                ws[mfg_cell].fill = yellow_fill
                                ws[pn_cell].fill = yellow_fill
                                file_modified[0] = True
                                local_file_errors.append({
                                    "Row": f"{mfg_cell}/{pn_cell}", "Field": f"{row_label} MFG Info",
                                    "Expected": "One of: " + ", ".join(expected_display_list), 
                                    "Found": f"({final_found_mfg} / {final_found_pn})", "Status": "ERROR"
                                })

                    # Header Fields Check
                    check_cell("D2", "Header Part Number", parent_part)
                    check_cell("D3", "Header Revision", parent_rev)
                    check_cell("D4", "Header Description", str(parent_info['Description']).strip())
                    check_cell("D5", "VSE Revision", "00")
                    check_cell("D6", "Customer ID", "APP01")
                    
                    # Row 9 (Parent Context Data Check)
                    check_cell("A9", "Row 9 Level", str(parent_info['Level']).strip())
                    exp_seq = str(parent_info['Find No']).split('.')[0].zfill(4) if pd.notna(parent_info['Find No']) else "0000"
                    check_cell("B9", "Row 9 Find No", exp_seq)
                    check_cell("D9", "Parent Qty", str(parent_info['Quantity']).strip() if pd.notna(parent_info['Quantity']) else "1.0")
                    check_cell("E9", "Row 9 Part Number", parent_part)
                    check_cell("F9", "Row 9 Revision", parent_rev)
                    check_cell("G9", "Row 9 Description", str(parent_info['Description']).strip())
                    check_cell("I9", "Row 9 UOM", str(parent_info['UOM']).strip())
                    check_mfg_cells("J9", "K9", parent_part, "Row 9")

                    # Process Dynamic Children Rows
                    current_row = 10
                    for child in child_rows:
                        check_cell(f"A{current_row}", f"Row {current_row} Level", str(child['Level']).strip())
                        c_seq = str(child['Find No']).split('.')[0].zfill(4) if pd.notna(child['Find No']) else "0000"
                        check_cell(f"B{current_row}", f"Row {current_row} Find No", c_seq)
                        check_cell(f"D{current_row}", f"Child Qty", str(child['Quantity']).strip() if pd.notna(child['Quantity']) else "0.0")
                        c_part = str(child['Part Number']).strip()
                        check_cell(f"E{current_row}", f"Row {current_row} Part Number", c_part)
                        
                        c_rev = str(child['Revision']).strip() if pd.notna(child['Revision']) else ""
                        if c_rev.lower() == 'nan': c_rev = ""
                        elif len(c_rev) == 1 and c_rev.isdigit(): c_rev = c_rev.zfill(2)
                        
                        check_cell(f"F{current_row}", f"Row {current_row} Revision", c_rev)
                        check_cell(f"G{current_row}", f"Row {current_row} Description", str(child['Description']).strip())
                        check_cell(f"I{current_row}", f"Row {current_row} UOM", str(child['UOM']).strip())
                        check_mfg_cells(f"J{current_row}", f"K{current_row}", c_part, f"Row {current_row}")
                            
                        current_row += 1
                    
                    if len(local_file_errors) > 0:
                        has_any_errors = True
                        for err in local_file_errors:
                            unified_report_rows.append({
                                "File": expected_filename, "Audit Result": "FAILED (Data Mismatch)",
                                "Row": err["Row"], "Field": err["Field"], "Expected": err["Expected"],
                                "Found": err["Found"], "Status": err["Status"]
                            })
                    else:
                        unified_report_rows.append({
                            "File": expected_filename, "Audit Result": "PASSED (100% Match)",
                            "Row": "", "Field": "", "Expected": "", "Found": "", "Status": "CLEAN"
                        })
                        
                except Exception as e:
                    has_any_errors = True
                    unified_report_rows.append({
                        "File": expected_filename, "Audit Result": "CRITICAL ERROR", "Row": "All", 
                        "Field": "File Read Error", "Expected": "Should open properly", 
                        "Found": f"Error: {str(e)}", "Status": "CORRUPTED"
                    })

            # Output Dashboard Rendering Section
            st.subheader("📋 BAVA TECH Audit Unified Matrix Summary")
            df_master_report = pd.DataFrame(unified_report_rows)
            
            if not has_any_errors:
                st.html("""
                    <div style="background-color:#d4edda; border-left: 8px solid #28a745; padding: 20px; border-radius: 5px; margin-bottom: 20px;">
                        <h3 style="color:#155724; margin:0;"><b>🎉 BAVA TECH VERIFICATION SUCCESS!</b></h3>
                        <p style="color:#155724; margin:10px 0 0 0; font-size:16px;">
                            <b>All submitted production files match perfectly with the master Indented BOM structure. Zero defects tracked.</b>
                        </p>
                    </div>
                """)
            else:
                st.error("⚠️ Audit Matrix detected data mismatches or missing attachments across processing pipelines.")
                
            st.write("### Master Audit Trail Tracker:")
            st.dataframe(df_master_report, width='stretch', hide_index=True)
            
            # Build optimized Unified Master Excel Report
            report_buffer = io.BytesIO()
            with pd.ExcelWriter(report_buffer, engine='openpyxl') as writer:
                df_master_report.to_excel(writer, index=False, sheet_name='Master_Audit_Report')
                ws = writer.sheets['Master_Audit_Report']
                
                green_bold_font = Font(name="Calibri", size=11, bold=True, color="008000")
                red_bold_font = Font(name="Calibri", size=11, bold=True, color="FF0000")
                
                for row_idx in range(2, ws.max_row + 1):
                    audit_cell = ws.cell(row=row_idx, column=2)
                    status_cell = ws.cell(row=row_idx, column=7)
                    
                    if "PASSED" in str(audit_cell.value).upper() or "CLEAN" in str(status_cell.value).upper():
                        audit_cell.font = green_bold_font
                        status_cell.font = green_bold_font
                    elif "FAILED" in str(audit_cell.value).upper() or "ERROR" in str(status_cell.value).upper() or "CRITICAL" in str(status_cell.value).upper():
                        audit_cell.font = red_bold_font
                        status_cell.font = red_bold_font
                
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = col[0].column_letter
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
            
            st.download_button(
                label="📥 Download Unified Master Audit Report (Excel)",
                data=report_buffer.getvalue(),
                file_name="BAVATECH_Unified_BOM_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.info("💡 Please upload the Master BOM and component targets to execute the BAVA TECH Audit workflow.")
